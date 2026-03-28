from __future__ import annotations

import io
from datetime import date
from typing import Any

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from sqlalchemy.orm import Session

from core.database import get_db
from models.warehouse import Warehouse
from schemas.excel_import import ExcelImportResult
from schemas.stock_transaction import StockTransactionCreate
from services.import_resolve import (
    build_item_id_by_code,
    get_or_create_warehouse_by_label,
    resolve_default_warehouse,
    resolve_item_id,
)
from services.inventory_upsert import upsert_inventory_snapshot
from services.issuance_pivot_parser import parse_issuance_pivot_lines
from services.stock_movement import apply_stock_transaction
from services.wms_inventory_parser import parse_wms_header_indices, parse_wms_row


router = APIRouter(prefix="/api/imports/excel", tags=["excel_import"])

_MAX_ERRORS = 80


def _pick_sheet_wb(wb: Any, preferred: str | None, *candidates: str) -> tuple[Any, str]:
    if preferred and preferred in wb.sheetnames:
        return wb[preferred], preferred
    for name in wb.sheetnames:
        for c in candidates:
            if c in name:
                return wb[name], name
    name = wb.sheetnames[0]
    return wb[name], name


@router.post("/issuance-pivot", response_model=ExcelImportResult)
async def import_issuance_pivot_excel(
    file: UploadFile = File(...),
    warehouse_id: int | None = Query(
        None,
        description="미지정 시 code=MAIN 창고(없으면 첫 창고)",
    ),
    sheet_name: str | None = Query(None, description="시트명. 비우면 출고량 등 자동 탐색"),
    db: Session = Depends(get_db),
) -> ExcelImportResult:
    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="빈 파일입니다.")

    wb = None
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=False)
        ws, used_name = _pick_sheet_wb(wb, sheet_name, "출고량", "출고")
        rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"엑셀을 열 수 없습니다: {e}") from e
    finally:
        if wb is not None:
            wb.close()

    lines, parse_errors = parse_issuance_pivot_lines(rows)
    errors = list(parse_errors)

    if warehouse_id is not None:
        wh = db.query(Warehouse).filter(Warehouse.id == warehouse_id).first()
        if not wh:
            raise HTTPException(status_code=400, detail="warehouse_id 를 찾을 수 없습니다.")
    else:
        try:
            wh = resolve_default_warehouse(db)
        except ValueError:
            raise HTTPException(
                status_code=400,
                detail="창고가 없습니다. 창고를 등록하거나 warehouse_id를 지정하세요.",
            ) from None

    by_code = build_item_id_by_code(db)
    fn = (file.filename or "upload")[:200]
    created = 0
    for line in lines:
        iid = resolve_item_id(by_code, line.item_code)
        if iid is None:
            errors.append(f"미등록 품목코드: {line.item_code}")
            if len(errors) >= _MAX_ERRORS:
                break
            continue
        ref = f"{fn}|출고량|{line.item_code}|{line.as_of_date.isoformat()}"
        if len(ref) > 450:
            ref = ref[:450]
        try:
            apply_stock_transaction(
                db,
                StockTransactionCreate(
                    item_id=iid,
                    warehouse_id=wh.id,
                    trx_type="OUT",
                    qty=line.qty,
                    as_of_date=line.as_of_date,
                    source_ref=ref,
                    reason="출고량 시트 일자별 출고 import",
                ),
            )
            created += 1
        except HTTPException as he:
            d = he.detail
            if isinstance(d, list):
                d = str(d)
            errors.append(f"{line.item_code} {line.as_of_date}: {d}")
            if len(errors) >= _MAX_ERRORS:
                break

    return ExcelImportResult(
        kind="issuance_pivot",
        sheet_used=used_name,
        rows_read=len(rows),
        transactions_created=created,
        errors=errors[:_MAX_ERRORS],
    )


@router.post("/wms-inventory", response_model=ExcelImportResult)
async def import_wms_inventory_excel(
    file: UploadFile = File(...),
    as_of_date: date = Query(..., description="스냅샷 기준일 (재고DB 덤프 기준)"),
    sheet_name: str | None = Query(None, description="시트명. 비우면 재고DB 등 자동 탐색"),
    max_rows: int = Query(
        100_000,
        ge=1,
        le=500_000,
        description="처리 행 상한 (대용량 방지)",
    ),
    db: Session = Depends(get_db),
) -> ExcelImportResult:
    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="빈 파일입니다.")

    wb = None
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
        ws, used_name = _pick_sheet_wb(wb, sheet_name, "재고DB", "재고")
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            raise HTTPException(status_code=400, detail="빈 시트입니다.") from None

        meta = parse_wms_header_indices(tuple(header or ()))
        if meta is None:
            raise HTTPException(
                status_code=400,
                detail="재고DB 헤더(물류센터·품목·셀명·가용수량)를 인식하지 못했습니다.",
            )

        by_code = build_item_id_by_code(db)
        errors: list[str] = []
        upserted = 0
        processed = 0
        batch = 0

        try:
            for row in rows_iter:
                if processed >= max_rows:
                    errors.append(f"max_rows={max_rows} 에 도달하여 중단했습니다.")
                    break
                excel_row = processed + 2
                tup = tuple(row or ())
                line, err = parse_wms_row(meta, tup, excel_row)
                if err:
                    errors.append(err)
                    if len(errors) >= _MAX_ERRORS:
                        break
                    processed += 1
                    continue
                if not line:
                    processed += 1
                    continue
                iid = resolve_item_id(by_code, line.item_code)
                if iid is None:
                    errors.append(f"{excel_row}행 미등록 품목: {line.item_code}")
                    if len(errors) >= _MAX_ERRORS:
                        break
                    processed += 1
                    continue
                try:
                    wh = get_or_create_warehouse_by_label(db, line.hub_label)
                except ValueError:
                    errors.append(f"{excel_row}행 물류센터 값 없음")
                    if len(errors) >= _MAX_ERRORS:
                        break
                    processed += 1
                    continue

                upsert_inventory_snapshot(
                    db,
                    item_id=iid,
                    warehouse_id=wh.id,
                    as_of_date=as_of_date,
                    lot_no=line.lot_label,
                    qty=line.qty,
                    expiry_date=None,
                )
                upserted += 1
                batch += 1
                if batch >= 500:
                    db.flush()
                    batch = 0
                processed += 1

            db.commit()
        except Exception:
            db.rollback()
            raise
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"엑셀을 열 수 없습니다: {e}") from e
    finally:
        if wb is not None:
            wb.close()

    return ExcelImportResult(
        kind="wms_inventory",
        sheet_used=used_name,
        rows_read=processed + 1,
        inventory_upserted=upserted,
        errors=errors[:_MAX_ERRORS],
    )
